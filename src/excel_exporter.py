import openpyxl
from openpyxl.styles import PatternFill, Font
from src.config import OUTPUT_DIR

def export_traceability_matrix(analysis_results: list, filename: str = "traceability_report.xlsx"):
    """Writes validation results directly into a formatted Excel file."""
    filepath = OUTPUT_DIR / filename
    
    # Create workbook and select active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "E-E Traceability Report"
    
    # Define Headers
    headers = ["Requirement ID", "Component", "Validation Status", "LLM Analysis Notes"]
    ws.append(headers)
    
    # Style Headers
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        
    # Append Data
    for result in analysis_results:
        row = [
            result.get("req_id", "N/A"),
            result.get("component", "N/A"),
            result.get("status", "Unknown"),
            result.get("notes", "")
        ]
        ws.append(row)
        
    # Apply conditional formatting to Status column
    green_fill = PatternFill(start_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", fill_type="solid")
    
    for row in range(2, ws.max_row + 1):
        status_cell = ws.cell(row=row, column=3)
        if status_cell.value == "VALID":
            status_cell.fill = green_fill
        elif status_cell.value == "CONFLICT":
            status_cell.fill = red_fill
            
    wb.save(filepath)
    return str(filepath)